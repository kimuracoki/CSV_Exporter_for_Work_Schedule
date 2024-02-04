import openpyxl as excel
import csv
import datetime

def main():
    """
    エントリポイント
    """
    schedule_by_therapist = load_schedule_by_therapist()
    work_codes = load_work_codes()
    csv_header = ['Subject', 'Start Date', 'Start Time', 'End Date', 'End Time', 'All Day Event', 'Description', 'Location']

    for therapist, schedule in schedule_by_therapist.items():
        with open(f'{therapist}.csv', 'w') as f:
            writer = csv.writer(f)
            writer.writerow(csv_header)
            for d, c in schedule.items():
                if c is not None:
                    start_time = next((data["start"] for data in work_codes if data["code"] == "A"), None)
                    end_time = next((data["end"] for data in work_codes if data["code"] == "A"), None)
                    writer.writerow(['Aman Grace', d, start_time, d, end_time, 'FALSE', '', ''])

def load_schedule_by_therapist():
    """
    シフト表を辞書型（を辞書型でネスト）で得る
    """
    work_schedule_book = excel.load_workbook('シフト表.xlsx', data_only=True)
    work_schedule_sheet = work_schedule_book['シフト表'] 

    days = []
    for column in work_schedule_sheet.iter_cols(min_col=11, max_col=11, min_row=6, max_row=36, values_only=True):
        for value in column:
            if value is not None:
                days.append(value.strftime('%Y/%m/%d'))
    
    therapists = {}
    for row in work_schedule_sheet.iter_rows(min_col=12, max_col=30, min_row=5, max_row=5, values_only=True):
        for value in row:
            if value is not None:
                therapists[value] = {}

    for i, t in enumerate(therapists):
        schedule = {}
        for j in range(31):
            schedule[days[j]] = work_schedule_sheet.cell(row=j+6, column=i+12).value
        therapists[t] = schedule

    work_schedule_book.close() 

    return therapists

def load_work_codes():
    """
    勤務コードを辞書型のリストで得る
    """
    work_schedule_book = excel.load_workbook('シフト表.xlsx', data_only=True)
    work_schedule_sheet = work_schedule_book['勤務コード定義']  

    codes = ['code', 'start', 'end', 'shop']

    work_codes = []
    for i in range(15): # TODOx: コードのMax値は暫定で15。設定ファイルで管理したい
        work_code = {}
        for j, c in enumerate(codes):
            value = work_schedule_sheet.cell(row=i+5, column=j+2).value
            if value is not None:
                if isinstance(value, datetime.time):
                    value = value.strftime('%H:%M')
                work_code[c] = value
        work_codes.append(work_code)

    work_schedule_book.close() 

    return work_codes

if __name__ == "__main__":
    main() 